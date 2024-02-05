import os
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from orquesta_sdk import Orquesta, OrquestaClientOptions
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# Initialize FastAPI app
app = FastAPI()

# Set up CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Initialize Orquesta client
def init_orquesta_client():
    api_key = os.getenv("ORQUESTA_API_KEY")
    options = OrquestaClientOptions(api_key=api_key, environment="production")
    return Orquesta(options)

client = init_orquesta_client()

initial_questions_with_options = [
    # Question 1: Conditional based on answer to question in the static front-end
    ("1", "Wat is uw relatie tot die ander?", ["ouder/verzorger", "echtgeno(o)t(e)/partner", "(schoon)zoon/(schoon)dochter", "mantelzorger/verzorger/familielid"], "1=ander"),
    # Question 2
    ("2", "Heeft u voldoende tijd (maximaal 10 minuten) om een aantal vragen over uw klacht te beantwoorden?", ["ja", "nee"], None),
    # Question 3
    ("3", "Op welk van de volgende gebieden heeft uw klacht betrekking? (er zijn meerdere antwoorden mogelijk)", ["stem", "keel", "spraak", "niet vloeiend spreken", "taal", "slikken", "adem", "gehoor", "mondgewoonten", "neurologisch probleem", "oncologisch probleem", "psychisch/psychiatrisch probleem", "leer-/ontwikkelingsprobleem", "anders"], None),
    # Question 4: Conditional based on answers to question 4
    ("4", "Is er door uw huisarts of specialist een diagnose gesteld?", ["ja", "nee"], "4=neurologisch probleem,oncologisch probleem,psychisch/psychiatrisch probleem,leer-/ontwikkelingsprobleem,anders"),
    # Question 5: Conditional based on answer to question 5 being "ja"
    ("5", "Hoe luidde die diagnose?", [], "5=ja")
]

# Function to load initial questions
def load_initial_questions():
    return initial_questions_with_options

def load_questions_from_sheet(sheet_path):
    questions_with_options = []
    workbook = load_workbook(sheet_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=3, values_only=True):
        question_index = row[0]  # Question index including subletters
        question = row[1]
        quick_reply_options = row[2].split(',') if row[2] else []
        condition = row[3] if len(row) > 3 else None
        if question:
            questions_with_options.append((question_index, question, quick_reply_options, condition))
    return questions_with_options

all_questions_with_options = load_initial_questions() + load_questions_from_sheet("data/vragenlijst.xlsx")


# Load the questions and options when the app starts
questions_with_options = load_questions_from_sheet("data/vragenlijst.xlsx")
@app.post("/question/")
async def question(request: Request):
    data = await request.json()
    question_index = data.get("question_index")
    previous_question = data.get("previous_question")
    previous_answer = data.get("previous_answer")

    # Log the incoming request data
    print(f"Received question_index={question_index}, previous_answer='{previous_answer}'")

    # Combine initial questions with main questions for seamless transition
    combined_questions_with_options = all_questions_with_options

    if question_index is None or question_index < 1:
        raise HTTPException(status_code=400, detail="Invalid question index")

    # Log the start of the condition checking loop
    print(f"Starting condition checking loop for question_index={question_index}")

    while question_index <= len(combined_questions_with_options):
        q_index, question, quick_reply_options, condition = combined_questions_with_options[question_index - 1]

        # Log current question and condition
        print(f"Checking question {q_index} with condition '{condition}'")

        if condition:
            condition_met = is_condition_met(condition, previous_answer, combined_questions_with_options)
            # Log the result of the condition check
            print(f"Condition '{condition}' met: {condition_met}")
            if not condition_met:
                question_index += 1
                # Log the increment of question_index if condition not met
                print(f"Condition not met, skipping to question_index={question_index}")
                continue

        # Log the question being returned
        print(f"Condition met or no condition for question {q_index}, proceeding to return this question.")
        break

    if question_index > len(combined_questions_with_options):
        raise HTTPException(status_code=400, detail="No suitable question found")

    # Assuming previous context logic remains the same
    previous_context = f"Vraag: {previous_question}\nAntwoord: {previous_answer}" if previous_question and previous_answer else ""

    # Assuming deployment logic remains the same
    deployment = client.deployments.invoke(
        key="logopedica-vragenlijsten",
        context={"environments": [], "klacht": ["mondgewoonten"]},
        inputs={"question": question, "previous": previous_context}
    )
    rephrased_question = deployment.choices[0].message.content

    # Log the question being sent to the frontend
    print(f"Sending to frontend: rephrased_question='{rephrased_question}', quick_reply_options={quick_reply_options}")

    return {"rephrased_question": rephrased_question, "quick_reply_options": quick_reply_options}

def is_condition_met(condition, previous_answer, combined_questions_with_options):
    condition_question_index, valid_answers = condition.split('=')
    # Split the valid answers by comma to support multiple valid answers
    valid_answers_list = valid_answers.split(',')
    # No need to adjust for zero-based indexing if your question indices already start at 1
    return previous_answer in valid_answers_list

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
